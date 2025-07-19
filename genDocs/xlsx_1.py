import io
import openpyxl
from flask import send_file
from datetime import datetime, date
from openpyxl import Workbook, load_workbook

DESTINY_PATH = "static/docs/"


def reporteCobrosXlsx(clientes, ventas, createfile):
    formatted_date = datetime.today().date()
    wb = Workbook()
    ws = wb.active
    ws.title = createfile
    ws.append([''])

    for cliente in clientes:
        ws.append([cliente['nombre']])
        for venta in ventas:
            if venta['cliente'] == cliente['nombre']:
                ws.append([venta['qty'], venta['producto'], venta['total']])

        ws.append([''])
        ws.append([''])
        ws.append(['', 'Saldo Anterior:', cliente['deuda']])
        ws.append(['Rosa Liliana Rivera, C.C. 55.177.973'])
        ws.append(['C. Ahorros No. 18876368419. Bancolombia'])
        ws.append(['Nequi: 310 340 2169'])
        ws.append(['Pago:', '', ])
        ws.append([''])

    wb.save(DESTINY_PATH + createfile + str(formatted_date) + '.xlsx')

    return


def reporteCarteraXlsx(allnames, cartera, totcartera, createfile):
    formatted_date = datetime.today().date()
    wb = Workbook()
    ws = wb.active
    ws.title = createfile

    for name in allnames:
        ws.append([''])
        ws.append(['', name['nombre']])
        ws.append([''])
        ws.append(['Id', 'Descripcion', 'Revista', 'Cant', 'Venta', 'Abono', 'FECHA'])

        for item in cartera:
            if name['nombre'] == item['cliente']:
                ws.append([item["id"], item['descripcion'], item['revista'], item['qty'], item['Vtotal'], item['Atotal'], item['fecha']])

        for totales in totcartera:
            if name['nombre'] == totales['cliente']:
                ws.append(['', '', 'TOTALES', '', totales['totVentas'], totales['totAbonos'], ''])
                ws.append(['', '', 'GrandTotal', '', '',  totales['grandTot'], ''])

    ws.append([''])
    ws.append([''])
    wb.save(DESTINY_PATH + createfile + str(formatted_date) + '.xlsx')

    return


def createXlsx(InventoryUpdate):
    formatted_date = datetime.today().date()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    for row in InventoryUpdate:
        ws.append(row)

    wb.save(DESTINY_PATH + 'Inventario_' + str(formatted_date) + '.xlsx')

    return

